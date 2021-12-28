from os import path, mkdir
from shutil import copyfile
from openpyxl import load_workbook
from tkinter import messagebox, Toplevel, StringVar, Text
from tkinter.ttk import LabelFrame, Frame, Button
from datetime import datetime


from general_funcs import GetVar


class CreateNewProjectnAskDesc:
    def __init__(self):
        super().__init__()

        tracking_stats_loc = 'Python_source\\!working_files\\dws_tracking_vals.txt'
        tracking_stats_all = GetVar(tracking_stats_loc, True)
        tracking_comment_line = tracking_stats_all[0]
        self.tracking_stats = tracking_stats_all[1]

        self.projects_t_date = self.tracking_stats[0]


        self.temp_delete_win = Toplevel()
        self.temp_delete_win.title('New Project')
        temp_d_frame = Frame(self.temp_delete_win)
        temp_d_frame.grid()

        entry_lf = LabelFrame(temp_d_frame, text="Project Description")
        entry_lf.grid(row=0)


        self.descript_text = Text(entry_lf, height=3, width=30)
        self.descript_text.grid()

        sub_butt = Button(
            temp_d_frame,
            text="Create Project",
            command=self.submitProjDesc
        )
        sub_butt.grid(row=1)

    def submitProjDesc(self, event=None):
        entry_got_got = self.descript_text.get("1.0", "end")

        # getting date info
        now_is_time = datetime.now()
        date_string = now_is_time.strftime('%m%y')
        month_int = now_is_time.strftime('%m')

        prev_proj_mont = self.tracking_stats[1]

        if month_int == '01' and (prev_proj_mont == '12' or prev_proj_mont == '11'):
            self.projects_t_date = '01'

        new_project_num = date_string + self.projects_t_date


        if len(entry_got_got) == 1:
            error_title = f'Description Empty'
            error_message = f'Do you want "{new_project_num}" to have an empty description?\n' \
                      f'Press "Yes" to Continue\n' \
                      f'Press "No" to Cancel'

            empty_q = messagebox.askyesno(title=error_title, message=error_message)

            if not empty_q:
                return

        title = f'New Project w/{new_project_num}?'
        message = f'Do you want "{new_project_num}" to be the Project Number?\n' \
                  f'Description: {entry_got_got}\n' \
                  f'Press "Yes" to Continue\n' \
                  f'Press "No" to Cancel'



        yes = messagebox.askyesno(title=title, message=message)

        if not yes:
            self.temp_delete_win.destroy()
            return
        if yes:
            CreateNewProject(descript=entry_got_got, top_level=self.temp_delete_win)






def CreateNewProject(shop_q=False, descript=None, top_level=None):
    # getting variables
    # new_proj_loc_var = GetVar('Python_Source\\!variables\\new_project_var.txt', False)
    d2_loc = GetVar('Python_Source\\!variables\\Quality_Control_Files\\QC-D2-Locations.txt', False)
    new_proj_loc_var = GetVar('Python_Source\\!variables\\new_project_var.txt', edit_q=False)

    tracking_stats_loc = 'Python_source\\!working_files\\dws_tracking_vals.txt'
    tracking_stats_all = GetVar(tracking_stats_loc, True)
    tracking_comment_line = tracking_stats_all[0]
    tracking_stats = tracking_stats_all[1]
    tracking_alllines = tracking_stats_all[-1]

    # print(new_proj_loc_var)
    print(tracking_stats)

    # xlsx templates to copy
    # Python_Source\!working_files\Template.xlsx
    xlsx_template_loc = new_proj_loc_var[0]
    text_stats_temp_loc = new_proj_loc_var[1]
    # Quality_Control\\!D - Documents\\D2 - Documentation\\D2-7 - Packing Slip.xlsx
    packingslip_template_loc = d2_loc[6]


    # getting amount of previously created projects
    projects_t_date = tracking_stats[0]

    # getting the month of the previous project creation to automatically start the count from 0 in the next year
    prev_proj_mont = tracking_stats[1]

    # getting date info
    now_is_time = datetime.now()
    date_string = now_is_time.strftime('%m%y')
    month_int = now_is_time.strftime('%m')
    month_str = now_is_time.strftime('%B')
    print(f'Month: {month_int}')
    if month_int == '01' and (prev_proj_mont == '12' or prev_proj_mont == '11'):
        projects_t_date = '01'

    new_project_num = date_string + projects_t_date

    if shop_q:
        title = f'New Month {month_str}'
        message = f'Creating New Month Project {month_str}'

        messagebox.showinfo(title=title, message=message)

        year = now_is_time.year
        month = now_is_time.strftime('%B')

        # project folder locations
        year_folder_path = f'Shop\\{year}'
        folder_path = f"{year_folder_path}\\{month}"
        project_pathxlsx = f'{folder_path}\\{month}-master.xlsx'
        src_folder = f'{folder_path}\\!src'
        purchase_folder = f'{folder_path}\\Purchase_Scans'
        certs_folder = f'{folder_path}\\Material_Cert_Scans'
        general_scans = f'{folder_path}\\General_Scans'

        if not path.isdir(year_folder_path):
            mkdir(year_folder_path)

        # making new project folder
        mkdir(folder_path)

        project = month
    else:

        new_num = int(projects_t_date) + 1

        project = new_project_num

        folder_path = f'Projects\\{project}'

        if path.isdir(folder_path):
            warning_message = f'Error Creating Project: {project}\n' \
                              f'Project {folder_path} already exists'
            messagebox.showwarning(title='Cannot Create Project', message=warning_message)
            return

        with open(tracking_stats_loc, 'w') as file:
            tracking_alllines[0+int(tracking_comment_line)] = f'0Projects Created This Year: {new_num:02d}\n'
            tracking_alllines[1+int(tracking_comment_line)] = f'1Month of Last Create Project: {month_int}\n'
            file.writelines(tracking_alllines)



        # project folder locations

        project_pathxlsx = f'{folder_path}\\{project}-master.xlsx'
        src_folder = f'{folder_path}\\!src'
        purchase_folder = f'{folder_path}\\Purchase_Scans'
        certs_folder = f'{folder_path}\\Material_Cert_Scans'
        general_scans = f'{folder_path}\\General_Scans'

        ini_loc =f"{folder_path}\\desktop.ini"
        input_lines = ["[{F29F85E0-4FF9-1068-AB91-08002B27B3D9}]\n", f"Prop5=31,{descript}"]

        # making new project folder
        mkdir(folder_path)

        with open(ini_loc, "w") as ini_file:
            ini_file.writelines(input_lines)

        top_level.destroy()







    # copying the file
    copyfile(xlsx_template_loc, project_pathxlsx)
    print('Creating {}.xlsx @ {}, yo'.format(project, project_pathxlsx))

    # opening workbook
    ss = load_workbook(project_pathxlsx)

    # new names of sheets
    sheet_name_lst = [
        project + "_dataSheet",
        project + "_purchases",
        project + "_financing"
    ]

    j = 0

    sheet_lst = ss.sheetnames
    # print("List of Template Sheets: {}".format(sheet_lst))
    # renaming the sheets
    for sheet in sheet_lst:
        # print("Sheet Renaming Cycle: {}. Should be {}.".format(i + 1, len(sheet_lst)))
        # declaring what sheet to rename
        ss_sheet = ss[sheet]
        # creating default sheet name variable
        sheet_name = sheet_name_lst[j]
        # renaming sheet
        ss_sheet.title = sheet_name
        j += 1
    print('Sheets Renamed')
    # saving sheet
    ss.save(project_pathxlsx)

    # # copying the file
    # copyfile(packingslip_template_loc, packingslip_pathxlsx)
    # print(f"Made {project} Packing Slip Folder")
    # AddDataToExcel(
    #     excel_loc=packingslip_pathxlsx,
    #     sheet_name='INPUT',
    #     col_loc=[1],
    #     row_list_data=[str(project)],
    #     place_loc=(0, 0),
    #     scan_max=(2, 2)
    # )

    # making source file
    mkdir(src_folder)
    print(f"Made {project} src Folder")

    # making the project stats file to check if packing slips or certain purchases have been made
    stats_text_loc = f'{src_folder}\\stats.txt'
    copyfile(text_stats_temp_loc, stats_text_loc)
    print(f"Made {project} Stats File")

    # making general/other scan folder
    mkdir(general_scans)
    print(f"Made {project} General_Scans Folder")

    # making purchase scan folder
    mkdir(purchase_folder)
    print(f"Made {project} Purchase_Scans Folder")

    # making cert scan folder
    mkdir(certs_folder)
    print(f"Made {project} Material_Cert_Scans Folder")

    return project
